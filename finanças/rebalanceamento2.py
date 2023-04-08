import openpyxl
from openpyxl.chart import BarChart, Reference

# cria uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active

# adiciona o cabeçalho
ws['A1'] = 'Ativo'
ws['B1'] = 'Quantidade'
ws['C1'] = 'Preço'
ws['D1'] = 'Porcentagem desejada'
ws['E1'] = 'Porcentagem atual'
ws['F1'] = 'Quantidade adicional'

# adiciona os ativos
row = 2
ativos = []

while True:
    nome = input("Insira o nome do ativo (ou 'fim' para terminar): ")
    if nome == 'fim':
        break
    quantidade = float(input("Insira a quantidade atual do ativo: "))
    preco = float(input("Insira o preço atual do ativo: "))
    porcentagem = float(input(f"Insira a porcentagem desejada para {nome}: "))
    porcentagem_atual = 0.0
    quantidade_adicional = 0.0
    
    ativos.append({'nome': nome, 'quantidade': quantidade, 'preco': preco, 
                   'porcentagem': porcentagem, 'porcentagem_atual': porcentagem_atual,
                   'quantidade_adicional': quantidade_adicional})
    
    # adiciona o ativo na planilha
    ws[f'A{row}'] = nome
    ws[f'B{row}'] = quantidade
    ws[f'C{row}'] = preco
    ws[f'D{row}'] = porcentagem
    
    row += 1

# calcula as porcentagens atuais e as quantidades adicionais necessárias
valor_total = sum([ativo['quantidade'] * ativo['preco'] for ativo in ativos])
for ativo in ativos:
    porcentagem_atual = (ativo['quantidade'] * ativo['preco'] / valor_total) * 100
    ativo['porcentagem_atual'] = porcentagem_atual
    if porcentagem_atual < ativo['porcentagem']:
        porcentagem_desejada = ativo['porcentagem']
        quantidade_atual = ativo['quantidade']
        preco_atual = ativo['preco']
        quantidade_adicional = (porcentagem_desejada / 100) * valor_total - quantidade_atual * preco_atual
        ativo['quantidade_adicional'] = quantidade_adicional
    else:
        ativo['quantidade_adicional'] = 0

# adiciona as informações na planilha
row = 2
for ativo in ativos:
    ws[f'E{row}'] = ativo['porcentagem_atual']
    ws[f'F{row}'] = ativo['quantidade_adicional']
    row += 1

# adiciona os gráficos na planilha
chart1 = BarChart()
chart1.title = "Quantidade atual de ativos"
chart1.y_axis.title = "Quantidade"
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=row-1)
cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
chart1.add_data(data)
chart1.set_categories(cats)
ws.add_chart(chart1, "H2")

chart2 = BarChart()
chart2.title = "Quantidade de ativos desejada"
chart2.y_axis.title = "Quantidade"
data = Reference(ws, min_col=6, min_row=2, max_col=6, max_row=row-1)
chart2.add_data(data)
chart2.set_categories(cats)
ws.add_chart(chart2, "H20")
